"""Excel report: the workbook downloads with the expected headers + auto-filter,
and the download-filename scheme (full / partial / spanning months)."""
import io
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from tests.harness import standalone, has_openpyxl  # noqa: E402


def run(h):
    c = h.client

    h.section("report")
    r = c.get("/report/download?start=2000-01-01&end=2100-01-01")
    ok = r.status_code == 200 and len(r.data) > 0
    if ok and has_openpyxl():
        from openpyxl import load_workbook
        ws = load_workbook(io.BytesIO(r.data)).active
        headers = [ws.cell(row=1, column=i).value for i in range(1, 10)]
        ok = (headers[0] == "Work Date" and "Task Type" in headers and "Work Summary" in headers
              and "Billed Amount\n(USD $)" in headers and ws.auto_filter.ref == "A1:I1")
    h.check("report Excel built with correct headers + filter", ok)

    from timezone.views_reports import _report_filename
    h.check("report name: full month", _report_filename("2026-06-01", "2026-06-30") == "Hours_June'26")
    h.check("report name: partial month", _report_filename("2026-06-05", "2026-06-20") == "PARTIAL-Hours_June'26")
    h.check("report name: spanning months", _report_filename("2026-06-15", "2026-08-10") == "Hours_June'26-August'26")
    h.check("report name: client prefix",
            _report_filename("2026-06-01", "2026-06-30", "LorenCook") == "LorenCook_Hours_June'26")
    h.check("report name: client prefix strips spaces/unsafe chars",
            _report_filename("2026-06-01", "2026-06-30", "Another Client!") == "AnotherClient_Hours_June'26")


if __name__ == "__main__":
    standalone(run)

"""Excel report routes."""
import calendar
import io
import re
from datetime import date, datetime

from flask import (
    render_template, request, redirect, url_for, flash, send_file, abort,
)

from timezone import app
from timezone.config import *          # noqa: F401,F403
from timezone.database import get_db
from timezone.services import *        # noqa: F401,F403


def _report_filename(start, end, client=None):
    """Build the download name from the date range (2-digit year), optionally
    prefixed with the sanitised client name (e.g. 'LorenCook_'):
      full calendar month   -> [Client_]Hours_June'26
      partial single month  -> [Client_]PARTIAL-Hours_June'26
      spanning months        -> [Client_]Hours_June'26-August'26
    """
    try:
        s = datetime.strptime(start, "%Y-%m-%d").date()
        e = datetime.strptime(end, "%Y-%m-%d").date()
        if (s.year, s.month) == (e.year, e.month):
            last_day = calendar.monthrange(s.year, s.month)[1]
            label = "%s'%s" % (s.strftime("%B"), s.strftime("%y"))
            full_month = (s.day == 1 and e.day == last_day)
            name = ("Hours_%s" % label) if full_month else ("PARTIAL-Hours_%s" % label)
        else:
            name = "Hours_%s'%s-%s'%s" % (s.strftime("%B"), s.strftime("%y"),
                                          e.strftime("%B"), e.strftime("%y"))
    except (TypeError, ValueError):
        name = "Hours_%s_to_%s" % (start, end)
    # client name prefix (spaces / unsafe chars stripped for the filename)
    prefix = re.sub(r"[^A-Za-z0-9._-]", "", client or "")
    return ("%s_%s" % (prefix, name)) if prefix else name

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    HAVE_OPENPYXL = True
except Exception:  # pragma: no cover
    HAVE_OPENPYXL = False



@app.route("/report")
def report():
    db = get_db()
    cid = current_client_id(db)
    charges = db.execute(
        "SELECT * FROM charge_types WHERE client_id = ? ORDER BY is_base DESC, name", (cid,)
    ).fetchall()
    today = date.today()
    default_start = today.replace(day=1).isoformat()
    default_end = today.isoformat()
    return render_template(
        "report.html",
        charges=charges,
        default_start=default_start,
        default_end=default_end,
        have_openpyxl=HAVE_OPENPYXL,
    )


def build_hours_report_xlsx(db, cid, start, end):
    """Build the hours timesheet Excel report for one client + date range and
    return the .xlsx bytes. Shared by the /report download and the invoice email,
    which attaches it for the invoice's billed period. (Caller guarantees openpyxl
    is available — see HAVE_OPENPYXL.)"""
    rates = charge_rate_map(db, cid)
    rows = db.execute(
        "SELECT e.work_date, e.task_id, t.description AS task_desc, e.sub_task, "
        "e.description, e.hours, e.charge_method "
        "FROM timesheet_entries e LEFT JOIN tasks t "
        "ON t.client_id = e.client_id AND t.task_id = e.task_id "
        "WHERE e.client_id = ? AND e.work_date >= ? AND e.work_date <= ? "
        "ORDER BY e.work_date, e.task_id, e.id",
        (cid, start, end),
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    headers = ["Work Date", "Task ID", "Task Description", "Task Type", "Work Summary",
               "Hours", "Charge\nMethod", "Charge Rate\n(USD $)", "Billed Amount\n(USD $)"]
    SUMMARY_COL = 5            # Work Summary — fixed wide + wrap, not auto-sized
    ncols = len(headers)

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")   # thicker date separator
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    top_sep = Border(left=thin, right=thin, bottom=thin, top=medium)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    shade_fill = PatternFill("solid", fgColor=REPORT_SHADE)

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r = 2
    total_hours = 0.0
    total_amount = 0.0
    date_idx = -1
    i = 0
    while i < len(rows):
        date_val = rows[i]["work_date"]
        date_idx += 1
        shade = (date_idx % 2 == 1)
        date_start_r = r
        j = i
        # within a date, merge Task ID + Description per task group
        while j < len(rows) and rows[j]["work_date"] == date_val:
            task_val = rows[j]["task_id"]
            task_start_r = r
            k = j
            while (k < len(rows) and rows[k]["work_date"] == date_val
                   and rows[k]["task_id"] == task_val):
                rk = rows[k]
                rate = rates.get(rk["charge_method"], 0.0)
                amount = (rk["hours"] or 0) * rate
                total_hours += rk["hours"] or 0
                total_amount += amount
                # normalise Windows CRLF to a single LF so Excel doesn't render
                # a blank line between each typed line in wrapped text cells
                def _lf(s):
                    return (s or "").replace("\r\n", "\n").replace("\r", "\n")
                values = [
                    rk["work_date"], rk["task_id"], _lf(rk["task_desc"]),
                    rk["sub_task"] or "", _lf(rk["description"]),
                    rk["hours"] or 0, rk["charge_method"] or "", rate, amount,
                ]
                for c, v in enumerate(values, start=1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.border = border
                    if c == 6:
                        cell.number_format = "0.00"
                    if c in (8, 9):
                        cell.number_format = "#,##0.00"
                    # everything vertical-centered; long text columns wrap so they
                    # don't stretch the sheet, money right-aligned, rest centered
                    if c in (3, SUMMARY_COL):     # Task Description, Work Summary (prose)
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    elif c in (2, 4):             # Task ID, Task Type (wrap if long)
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    elif c == 9:                  # Billed Amount (currency)
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:                         # Work Date, Hours, Charge Method, Charge Rate
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    if shade:
                        cell.fill = shade_fill
                r += 1
                k += 1
            if r - 1 > task_start_r:
                ws.merge_cells(start_row=task_start_r, start_column=2, end_row=r - 1, end_column=2)
                ws.merge_cells(start_row=task_start_r, start_column=3, end_row=r - 1, end_column=3)
            j = k
        date_end_r = r - 1
        if date_end_r > date_start_r:
            ws.merge_cells(start_row=date_start_r, start_column=1, end_row=date_end_r, end_column=1)
        # thicker separator line above each date band
        for c in range(1, ncols + 1):
            ws.cell(row=date_start_r, column=c).border = top_sep
        i = j

    # TOTAL row (separated with a thick line)
    total_font = Font(bold=True)
    tcell = ws.cell(row=r, column=SUMMARY_COL, value="TOTAL")
    tcell.font = total_font
    tcell.alignment = Alignment(horizontal="right", vertical="center")
    hc = ws.cell(row=r, column=6, value=round(total_hours, 2))
    hc.font = total_font
    hc.number_format = "0.00"
    hc.alignment = Alignment(horizontal="center", vertical="center")
    ac = ws.cell(row=r, column=9, value=round(total_amount, 2))
    ac.font = total_font
    ac.number_format = "#,##0.00"
    ac.alignment = Alignment(horizontal="right", vertical="center")
    for c in range(1, ncols + 1):
        ws.cell(row=r, column=c).border = top_sep

    # add the header auto-filter + freeze BEFORE sizing, so the column widths
    # account for the filter dropdown buttons on the header row
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s1" % get_column_letter(ncols)

    # Column widths. Prose columns get a fixed width and wrap instead of
    # stretching to fit their longest value; the rest auto-size to content
    # (capped, so long Task IDs / Task Types wrap rather than widen the sheet).
    FIXED_WIDTHS = {3: 30, SUMMARY_COL: 55}   # Task Description, Work Summary
    CAP_WIDTHS = {2: 20, 4: 22}               # Task ID, Task Type
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        if c in FIXED_WIDTHS:
            ws.column_dimensions[letter].width = FIXED_WIDTHS[c]
            continue
        # widest line of the (possibly multi-line) header
        maxlen = max(len(line) for line in str(headers[c - 1]).split("\n"))
        for rr in range(2, r + 1):
            v = ws.cell(row=rr, column=c).value
            if v is None:
                continue
            s = ("%.2f" % v) if isinstance(v, float) else str(v)
            maxlen = max(maxlen, len(s))
        ws.column_dimensions[letter].width = min(maxlen + 2, CAP_WIDTHS.get(c, 40))

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


@app.route("/report/download")
def report_download():
    if not HAVE_OPENPYXL:
        abort(500, "openpyxl not installed")
    db = get_db()
    start = request.args.get("start")
    end = request.args.get("end")
    if not start or not end:
        flash("Please pick a date range.", "error")
        return redirect(url_for("report"))
    cid = current_client_id(db)
    data = build_hours_report_xlsx(db, cid, start, end)
    client = current_client(db)
    fname = _report_filename(start, end, client["name"] if client else None) + ".xlsx"
    return send_file(
        io.BytesIO(data), as_attachment=True, download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
